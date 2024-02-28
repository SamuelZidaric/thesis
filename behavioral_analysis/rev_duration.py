import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import argparse

def rev_duration(input_path, project_folder, frame_rate=80):
    # Loading the dataset
    data = pd.read_csv(input_path, header=None)
    data.columns = ['frames', 'values']

    # Identifying reversals in the data
    reversals = data['values'] == -1

    # Finding starts and ends of reversals in the data
    starts = data['frames'][reversals & (~reversals.shift(1, fill_value=False))]
    ends = data['frames'][reversals & (~reversals.shift(-1, fill_value=False))]

    reversal_stats = [(f"Reversal {i}", start, end, (end - start) / frame_rate)
                      for i, (start, end) in enumerate(zip(starts, ends), start=1)]

    # Creating a DataFrame for the reversal statistics
    reversal_stats_df = pd.DataFrame(reversal_stats, columns=["Reversal", "Start Frame", "End Frame", "Duration (seconds)"])

    # Creating the Excel file with centered alignment and autofit columns
    wb = openpyxl.Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(reversal_stats_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length

    # Save the workbook with a name matching the Snakemake expectation
    output_filename = "rev_duration_all.xlsx"
    output_path = os.path.join(project_folder, output_filename)

    # Save the workbook
    wb.save(output_path)
    print(f"Modified Excel file saved at: {output_path}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Process behavioral annotation data.')
    parser.add_argument('-i', '--input_path', help='Path to the behavioral annotation CSV file', required=True)
    args = vars(parser.parse_args())

    main_folder = args['input_path']
    project_folder = main_folder
    print("Project folder is: ", project_folder)

    input_path = os.path.join(project_folder, "beh_annotation1.csv")
    rev_duration(input_path, project_folder)

