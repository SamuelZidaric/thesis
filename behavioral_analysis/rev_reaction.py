import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import argparse

def rev_reaction(input_path, project_folder):
    # Load the data from the input file
    data = pd.read_csv(input_path, engine='python')

    # Constants for the analysis
    fps = 80  # frames per second
    rest_duration = 29  # seconds of resting
    first_rest_duration = 30  # seconds of resting for the first period
    activation_duration = 3  # seconds of led activation
    behavior_column = '0'

    # Calculate the number of frames for activation periods
    activation_frames = activation_duration * fps

    # Identifying the frame ranges for each activation period
    total_frames = len(data)
    activation_periods = []

    current_frame = 0
    first_period = True  # Indicator for the first resting period
    while current_frame < total_frames:
        # Adjust rest duration for the first period
        if first_period:
            rest_frames = first_rest_duration * fps
            first_period = False  # Reset the flag after the first period
        else:
            rest_frames = rest_duration * fps

        start_activation = current_frame + rest_frames
        end_activation = start_activation + activation_frames
        if end_activation > total_frames:
            end_activation = total_frames
        activation_periods.append((start_activation, end_activation))
        current_frame = end_activation

    # Check for reversals during these periods
    reversals_timing = []
    for i, (start, end) in enumerate(activation_periods):
        period_data = data.iloc[start:end]
        reversal_duration = None  # Initialize reversal duration

        if all(period_data[behavior_column] == 0):
            status = "No movement"
        elif period_data.iloc[0][behavior_column] == -1:
            status = "Already Reversing"
        elif any(period_data[behavior_column] == -1):
            first_reversal_frame = period_data[period_data[behavior_column] == -1].first_valid_index()
            time_to_reversal = (first_reversal_frame - start) / fps
            status = time_to_reversal

            # Find the end of the reversal beyond the current activation period
            extended_data = data.iloc[first_reversal_frame:]  # Start searching from the first reversal frame
            reversal_end_condition = extended_data[behavior_column] == 1
            if reversal_end_condition.any():
                reversal_end_frame = reversal_end_condition.idxmax()  # The first '1' after reversal starts
                reversal_duration = (reversal_end_frame - first_reversal_frame) / fps
            else:
                reversal_duration = (len(data) - first_reversal_frame) / fps  # Until the end if no '1' found
        else:
            status = "No Reversal"

        reversals_timing.append((i + 1, start, end, status, reversal_duration))    # Create the DataFrame
        
    reversal_data = pd.DataFrame(reversals_timing, columns=['Activation', 'Start Frame', 'End Frame', 'Status/Reaction Time', 'Reversal duration'])

    # Replacing None values with the string 'NaN'
    reversal_data_filled = reversal_data.fillna('NaN')

    # Creating the Excel file with centered alignment and autofit columns
    wb = openpyxl.Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(reversal_data_filled, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length

    # Save the workbook with a name matching the Snakemake expectation
    output_filename = "rev_reaction.xlsx"
    output_path = os.path.join(project_folder, output_filename)

    # Save the workbook
    wb.save(output_path)
    print(f"Modified Excel file saved at: {output_path}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Process behavioral annotation data.')
    parser.add_argument('-i', '--input_path', help='Path to the behavioral annotation CSV file', required=True)
    args = vars(parser.parse_args())

    main_folder = args['input_path']
    project_folder = main_folder
    print("Project folder is: ", project_folder)

    input_path = os.path.join(project_folder, "beh_annotation1.csv")
    rev_reaction(input_path, project_folder)
